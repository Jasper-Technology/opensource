"""Tests for the Excel template parser."""
from __future__ import annotations
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.import_ import parse_excel


def _make_workbook(
    *,
    blocks: list[dict] | None = None,
    streams: list[dict] | None = None,
    components: list[str] | None = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if components is not None:
        ws = wb.create_sheet("Components")
        ws.append(["name", "formula"])
        for c in components:
            ws.append([c, ""])

    if blocks is not None:
        ws = wb.create_sheet("Blocks")
        ws.append(["id", "type", "name", "duty", "duty_unit", "work", "work_unit"])
        for b in blocks:
            ws.append(
                [
                    b["id"],
                    b["type"],
                    b.get("name", b["id"]),
                    b.get("duty"),
                    b.get("duty_unit"),
                    b.get("work"),
                    b.get("work_unit"),
                ]
            )

    if streams is not None:
        ws = wb.create_sheet("Streams")
        base_headers = ["id", "from_block", "from_port", "to_block", "to_port", "T", "T_unit", "P", "P_unit", "flow", "flow_unit"]
        comp_headers = components or []
        ws.append(base_headers + comp_headers)
        for s in streams:
            base = [
                s["id"], s.get("from_block"), s.get("from_port"),
                s.get("to_block"), s.get("to_port"),
                s.get("T"), s.get("T_unit"),
                s.get("P"), s.get("P_unit"),
                s.get("flow"), s.get("flow_unit"),
            ]
            comp_vals = [s.get("comp", {}).get(c) for c in comp_headers]
            ws.append(base + comp_vals)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestBasicShape:
    def test_minimum_workbook(self) -> None:
        data = _make_workbook(
            blocks=[{"id": "P1", "type": "Pump", "work": 100, "work_unit": "kW"}],
        )
        result = parse_excel(data)
        assert result.project is not None
        assert len(result.project.blocks) == 1
        assert result.project.blocks[0].id == "P1"
        assert result.project.blocks[0].type == "Pump"
        # 100 kW → 100_000 W in SI
        assert result.project.blocks[0].work == pytest.approx(100_000)

    def test_missing_blocks_sheet_error(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Components").append(["name"])
        buf = BytesIO()
        wb.save(buf)
        result = parse_excel(buf.getvalue())
        assert result.has_errors
        assert result.project is None


class TestUnits:
    def test_converts_duty_to_si(self) -> None:
        data = _make_workbook(
            blocks=[{"id": "H1", "type": "Heater", "duty": 1, "duty_unit": "GJ/h"}],
        )
        result = parse_excel(data)
        duty = result.project.blocks[0].duty
        # 1 GJ/h = 1e9 J / 3600 s ≈ 277777.78 W
        assert duty == pytest.approx(1e9 / 3600, rel=1e-4)

    def test_si_unit_no_conversion(self) -> None:
        data = _make_workbook(
            blocks=[{"id": "H1", "type": "Heater", "duty": 5000, "duty_unit": "W"}],
        )
        assert parse_excel(data).project.blocks[0].duty == 5000

    def test_bad_unit_warns_not_fatal(self) -> None:
        data = _make_workbook(
            blocks=[{"id": "H1", "type": "Heater", "duty": 100, "duty_unit": "bogus"}],
        )
        result = parse_excel(data)
        assert result.project is not None
        assert result.count("warning") >= 1
        assert result.project.blocks[0].duty is None


class TestStreams:
    def test_parses_T_P_flow(self) -> None:
        data = _make_workbook(
            blocks=[{"id": "H1", "type": "Heater"}],
            streams=[
                {
                    "id": "S1",
                    "from_block": "Feed",
                    "from_port": "out",
                    "to_block": "H1",
                    "to_port": "in",
                    "T": 100,
                    "T_unit": "C",
                    "P": 5,
                    "P_unit": "bar",
                    "flow": 100,
                    "flow_unit": "kmol/h",
                }
            ],
        )
        result = parse_excel(data)
        s = result.project.streams[0]
        assert s.from_block == "Feed"
        assert s.T == pytest.approx(373.15)
        assert s.P == pytest.approx(5e5)
        assert s.flow_mol == pytest.approx(27.7778, rel=1e-3)

    def test_composition_columns(self) -> None:
        data = _make_workbook(
            components=["H2", "CO"],
            blocks=[{"id": "R1", "type": "Reactor"}],
            streams=[
                {
                    "id": "S1",
                    "to_block": "R1",
                    "T": 300, "T_unit": "K",
                    "P": 1, "P_unit": "bar",
                    "flow": 10, "flow_unit": "mol/s",
                    "comp": {"H2": 0.67, "CO": 0.33},
                }
            ],
        )
        result = parse_excel(data)
        assert result.project.streams[0].composition == {"H2": 0.67, "CO": 0.33}

    def test_mass_flow_unit(self) -> None:
        data = _make_workbook(
            blocks=[{"id": "P1", "type": "Pump"}],
            streams=[{"id": "S1", "to_block": "P1", "flow": 3600, "flow_unit": "kg/h"}],
        )
        s = parse_excel(data).project.streams[0]
        assert s.flow_mass == pytest.approx(1.0)
        assert s.flow_mol is None


class TestSecurity:
    def test_oversize_rejected(self) -> None:
        # Fabricate a pseudo-file over the 10 MB limit
        big = b"x" * (11 * 1024 * 1024)
        result = parse_excel(big)
        assert result.has_errors
        assert result.project is None

    def test_corrupt_workbook(self) -> None:
        result = parse_excel(b"not a zip file")
        assert result.has_errors
        assert result.project is None
