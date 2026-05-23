"""Tests for the Excel import endpoint."""
from __future__ import annotations
from io import BytesIO

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook


@pytest.fixture
def client() -> TestClient:
    import importlib
    route = importlib.import_module("app.api.routes.import_")
    app = FastAPI()
    app.include_router(route.router, prefix="/api")
    return TestClient(app)


def _simple_workbook() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    blocks = wb.create_sheet("Blocks")
    blocks.append(["id", "type", "name", "duty", "duty_unit", "work", "work_unit"])
    blocks.append(["P1", "Pump", "Feed Pump", None, None, 100, "kW"])
    blocks.append(["H1", "Heater", "Preheater", 1, "GJ/h", None, None])

    streams = wb.create_sheet("Streams")
    streams.append(["id", "from_block", "from_port", "to_block", "to_port", "T", "T_unit", "P", "P_unit", "flow", "flow_unit"])
    streams.append(["S1", "P1", "out", "H1", "in", 50, "C", 5, "bar", 10, "mol/s"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_excel_success(client: TestClient) -> None:
    data = _simple_workbook()
    r = client.post(
        "/api/import/excel",
        files={"file": ("project.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["project_name"] == "project"
    assert body["summary"]["blocks"] == 2
    assert body["summary"]["streams"] == 1
    assert body["summary"]["errors"] == 0

    # Units converted to SI on the wire
    pump = next(b for b in body["blocks"] if b["id"] == "P1")
    assert pump["work"] == pytest.approx(100_000)


def test_empty_file_rejected(client: TestClient) -> None:
    r = client.post(
        "/api/import/excel",
        files={"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400


def test_import_dwsim_success(client: TestClient) -> None:
    xml = b"""<?xml version="1.0"?>
<DWSIMSimulation><SimulationObjects>
  <SimulationObject><Name>P1</Name><ObjectType>Pump</ObjectType></SimulationObject>
  <SimulationObject>
    <Name>MS1</Name><ObjectType>MaterialStream</ObjectType>
    <DownstreamObject>P1</DownstreamObject>
    <Phase><Temperature>300</Temperature><Pressure>100000</Pressure><MolarFlow>10</MolarFlow></Phase>
  </SimulationObject>
</SimulationObjects></DWSIMSimulation>"""
    r = client.post(
        "/api/import/dwsim",
        files={"file": ("sim.dwxml", xml, "application/xml")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["blocks"] == 1
    assert body["summary"]["streams"] == 1


def test_import_describe_fallback_without_api_key(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    r = client.post(
        "/api/import/describe",
        json={"description": "methanol synthesis with a reactor and distillation column"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["blocks"] >= 3
    assert body["summary"]["errors"] == 0
    # Info diagnostic should mention the missing API key
    assert any("ANTHROPIC_API_KEY" in d["message"] for d in body["diagnostics"])


def test_import_describe_rejects_empty(client: TestClient) -> None:
    r = client.post("/api/import/describe", json={"description": ""})
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["errors"] >= 1
    assert body["summary"]["blocks"] == 0


def test_corrupt_file_returns_diagnostic(client: TestClient) -> None:
    r = client.post(
        "/api/import/excel",
        files={"file": ("bad.xlsx", b"not a real xlsx", "application/octet-stream")},
    )
    assert r.status_code == 200  # errors captured in body, not HTTP status
    body = r.json()
    assert body["summary"]["errors"] >= 1
    assert body["summary"]["blocks"] == 0
