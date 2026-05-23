"""Tests for POST /api/tea/run and GET /api/tea/catalog."""
from __future__ import annotations
import pytest
from fastapi.testclient import TestClient

from fastapi import FastAPI


@pytest.fixture
def client() -> TestClient:
    """Minimal app that mounts only the TEA router. Import tea directly to
    avoid triggering the routes package __init__ which imports simulate
    (pyomo-dependent)."""
    import importlib
    tea_route = importlib.import_module("app.api.routes.tea")
    app = FastAPI()
    app.include_router(tea_route.router, prefix="/api")
    return TestClient(app)


def test_catalog_returns_expected_classes(client: TestClient) -> None:
    r = client.get("/api/tea/catalog")
    assert r.status_code == 200
    data = r.json()
    assert "Pump" in data["equipment"]
    assert "HeatExchanger" in data["equipment"]
    assert data["cepci_reference"] > 0
    assert "Centrifugal" in data["equipment"]["Pump"]["subtypes"]


def test_run_simple_pump_project(client: TestClient) -> None:
    body = {
        "scenario": {
            "cepci": 800,
            "contingency": 0.18,
            "operating_hours": 8000,
            "default_material": "CarbonSteel",
            "discount_rate": 0.10,
            "project_life_years": 20,
            "annual_revenue": 5_000_000,
            "annual_feedstock_cost": 2_000_000,
        },
        "blocks": [
            {
                "id": "P1",
                "type": "Pump",
                "work": 100_000,  # 100 kW
                "inlet":  {"T": 300, "P": 1e5, "flow_mol": 10},
                "outlet": {"T": 300, "P": 10e5, "flow_mol": 10},
            },
            {
                "id": "H1",
                "type": "Heater",
                "duty": 1e6,
                "inlet":  {"T": 300, "P": 2e5, "flow_mol": 10},
                "outlet": {"T": 400, "P": 2e5, "flow_mol": 10},
            },
        ],
    }
    r = client.post("/api/tea/run", json=body)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["status"] == "ok"
    assert len(data["blocks"]) == 2

    # Both blocks costed
    for block in data["blocks"]:
        assert block["bare_module_usd"] is not None
        assert block["bare_module_usd"] > 0

    # Totals rollup
    assert data["totals"]["capex_total_module_usd"] > 0
    assert data["totals"]["opex_annual_usd"] > 0

    # Economics populated
    assert "npv_usd" in data["economics"]
    assert "irr" in data["economics"]
    assert "payback_years" in data["economics"]


def test_run_with_user_override(client: TestClient) -> None:
    body = {
        "scenario": {"cepci": 800, "default_material": "CarbonSteel"},
        "blocks": [
            {
                "id": "P1",
                "type": "Pump",
                "work": 100_000,
                "material": "StainlessSteel",
                "inlet": {"T": 300, "P": 1e5, "flow_mol": 10},
                "outlet": {"T": 300, "P": 10e5, "flow_mol": 10},
            },
        ],
    }
    r = client.post("/api/tea/run", json=body)
    assert r.status_code == 200
    block = r.json()["blocks"][0]
    assert block["material"] == "StainlessSteel"
    assert block["Fm"] == 2.31  # SS value for Centrifugal


def test_run_no_cost_block_returns_null(client: TestClient) -> None:
    body = {
        "scenario": {"cepci": 800},
        "blocks": [
            {"id": "Mixer1", "type": "Mixer"},
        ],
    }
    r = client.post("/api/tea/run", json=body)
    assert r.status_code == 200
    block = r.json()["blocks"][0]
    assert block["bare_module_usd"] is None
    assert block["note"] == "no-cost block"


def test_run_invalid_block_type_surfaces_error(client: TestClient) -> None:
    body = {
        "scenario": {"cepci": 800},
        "blocks": [{"id": "X1", "type": "Pump"}],  # missing work
    }
    r = client.post("/api/tea/run", json=body)
    # Status is still 200; per-block error captured in result
    assert r.status_code == 200
    assert r.json()["status"] == "partial"
    assert r.json()["blocks"][0]["error"]


def test_run_from_import_chains_streams(client: TestClient) -> None:
    """The import-chained endpoint should wire inlet/outlet streams by name."""
    body = {
        "scenario": {"cepci": 800, "annual_revenue": 1_000_000},
        "blocks": [
            {"id": "Feed", "type": "Feed"},
            {"id": "P1", "type": "Pump", "work": 50_000},
            {"id": "Sink", "type": "Sink"},
        ],
        "streams": [
            {
                "id": "S1", "from_block": "Feed", "from_port": "out",
                "to_block": "P1", "to_port": "in",
                "T": 300, "P": 1e5, "flow_mol": 10,
            },
            {
                "id": "S2", "from_block": "P1", "from_port": "out",
                "to_block": "Sink", "to_port": "in",
                "T": 300, "P": 5e5, "flow_mol": 10,
            },
        ],
    }
    r = client.post("/api/tea/run-from-import", json=body)
    assert r.status_code == 200, r.text
    data = r.json()
    # Pump should be costed; Feed and Sink are no-cost.
    pump = next(b for b in data["blocks"] if b["id"] == "P1")
    assert pump["bare_module_usd"] > 0
    assert pump["pressure_barg"] > 0  # 5 bar abs - 1 atm ≈ 4 barg


def test_export_xlsx_returns_workbook(client: TestClient) -> None:
    from io import BytesIO
    from openpyxl import load_workbook

    body = {
        "scenario": {"cepci": 800, "project_life_years": 20, "discount_rate": 0.10},
        "project_name": "MeOH retrofit",
        "result": {
            "status": "ok",
            "blocks": [],
            "totals": {
                "capex_bare_module_usd": 0,
                "capex_total_module_usd": 0,
                "opex_annual_usd": 0,
                "opex_utilities_usd": 0,
                "opex_labor_usd": 0,
                "opex_maintenance_usd": 0,
                "opex_feedstock_usd": 0,
                "revenue_annual_usd": 0,
            },
            "economics": {
                "npv_usd": 0, "irr": 0, "payback_years": -1, "tac_usd": 0, "lcop": -1,
            },
            "warnings": [],
        },
    }
    r = client.post("/api/tea/export/xlsx", json=body)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    # Slashes + spaces in project_name should be replaced with underscores
    assert "MeOH_retrofit.xlsx" in r.headers.get("content-disposition", "")
    wb = load_workbook(BytesIO(r.content))
    assert "Summary" in wb.sheetnames


def test_sensitivity_returns_sorted_drivers(client: TestClient) -> None:
    body = {
        "scenario": {
            "cepci": 800,
            "project_life_years": 20,
            "discount_rate": 0.10,
            "annual_revenue": 5_000_000,
            "annual_feedstock_cost": 1_500_000,
        },
        "blocks": [
            {"id": "F1", "type": "Feed"},
            {"id": "P1", "type": "Pump", "work": 50_000},
            {"id": "S1", "type": "Sink"},
        ],
        "streams": [
            {"id": "S1", "from_block": "F1", "to_block": "P1",
             "T": 300, "P": 1e5, "flow_mol": 10},
            {"id": "S2", "from_block": "P1", "to_block": "S1",
             "T": 300, "P": 5e5, "flow_mol": 10},
        ],
        "perturbation": 0.20,
    }
    r = client.post("/api/tea/sensitivity", json=body)
    assert r.status_code == 200, r.text
    data = r.json()
    assert "drivers" in data
    assert len(data["drivers"]) >= 3  # at least revenue + feedstock + CEPCI
    # Sorted by swing descending
    swings = [d["swing"] for d in data["drivers"]]
    assert swings == sorted(swings, reverse=True)


def test_monte_carlo_returns_histogram(client: TestClient) -> None:
    body = {
        "scenario": {
            "cepci": 800,
            "project_life_years": 20,
            "discount_rate": 0.10,
            "annual_revenue": 5_000_000,
            "annual_feedstock_cost": 1_500_000,
        },
        "blocks": [
            {"id": "F1", "type": "Feed"},
            {"id": "P1", "type": "Pump", "work": 50_000},
            {"id": "S1", "type": "Sink"},
        ],
        "streams": [
            {"id": "S1", "from_block": "F1", "to_block": "P1",
             "T": 300, "P": 1e5, "flow_mol": 10},
            {"id": "S2", "from_block": "P1", "to_block": "S1",
             "T": 300, "P": 5e5, "flow_mol": 10},
        ],
        "samples": 500,
        "seed": 42,
    }
    r = client.post("/api/tea/monte-carlo", json=body)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["samples"] == 500
    assert data["p10"] <= data["p50"] <= data["p90"]
    assert 0 <= data["probability_positive"] <= 1
    assert len(data["histogram"]) == 20
    assert sum(b["count"] for b in data["histogram"]) == 500


def test_esg_endpoint_aggregates_scope_1_2_3(client: TestClient) -> None:
    body = {
        "fuels": [{"fuel": "natural_gas", "mass_kg_per_year": 100_000}],
        "grid_region": "US-GC",
        "electricity_kwh_per_year": 500_000,
        "feedstocks": [{"name": "natural_gas", "mass_kg_per_year": 50_000}],
        "annual_production_kg": 1_000_000,
    }
    r = client.post("/api/tea/esg", json=body)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["scope_1_tco2_per_year"] > 0
    assert data["scope_2_tco2_per_year"] > 0
    assert data["scope_3_tco2_per_year"] > 0
    assert data["total_tco2_per_year"] == pytest.approx(
        data["scope_1_tco2_per_year"] + data["scope_2_tco2_per_year"]
        + data["scope_3_tco2_per_year"],
        rel=1e-6,
    )
    assert data["ci_kgco2_per_kg_product"] > 0


def test_tci_endpoint_rolls_up_direct_plus_indirect(client: TestClient) -> None:
    body = {
        "result": {
            "status": "ok",
            "warnings": [],
            "blocks": [
                {"id": "P1", "type": "Pump", "total_module_usd": 500_000,
                 "equipment_class": "Pump", "subtype": "Centrifugal",
                 "material": "CarbonSteel", "size": 100, "size_unit": "kW",
                 "pressure_barg": 4, "purchased_usd": 100_000,
                 "bare_module_usd": 400_000, "Fm": 1.54, "Fp": 1.0,
                 "in_range": True, "note": "", "error": None},
                {"id": "T1", "type": "Tank", "total_module_usd": 2_000_000,
                 "equipment_class": "Tank", "subtype": "APIFixedRoof",
                 "material": "CarbonSteel", "size": 1000, "size_unit": "m3",
                 "pressure_barg": 0, "purchased_usd": 500_000,
                 "bare_module_usd": 1_700_000, "Fm": 1.0, "Fp": 1.0,
                 "in_range": True, "note": "", "error": None},
            ],
            "totals": {
                "capex_bare_module_usd": 0,
                "capex_total_module_usd": 0,
                "opex_annual_usd": 0,
                "opex_utilities_usd": 0, "opex_labor_usd": 0,
                "opex_maintenance_usd": 0, "opex_feedstock_usd": 0,
                "revenue_annual_usd": 0,
            },
            "economics": {
                "npv_usd": 0, "irr": 0, "payback_years": -1,
                "tac_usd": 0, "lcop": -1,
            },
        },
        "annual_revenue": 5_000_000,
        "config": {"plant_type": "fluid", "osbl_ratio": 0.3},
    }
    r = client.post("/api/tea/tci", json=body)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["isbl_equipment"] == 500_000  # P1
    assert data["osbl_equipment"] == 2_000_000  # T1 (explicit OSBL, ratio not used)
    assert data["total_capital_investment"] > data["fixed_capital"]
    assert data["working_capital"] == pytest.approx(5_000_000 * 0.15)


def test_after_tax_endpoint_applies_macrs_and_credits(client: TestClient) -> None:
    body = {
        "scenario": {
            "cepci": 800, "project_life_years": 20, "discount_rate": 0.10,
        },
        "result": {
            "status": "ok", "blocks": [], "warnings": [],
            "totals": {
                "capex_bare_module_usd": 8_000_000,
                "capex_total_module_usd": 10_000_000,
                "opex_annual_usd": 2_000_000,
                "opex_utilities_usd": 0, "opex_labor_usd": 0,
                "opex_maintenance_usd": 0, "opex_feedstock_usd": 0,
                "revenue_annual_usd": 5_000_000,
            },
            "economics": {
                "npv_usd": 0, "irr": 0, "payback_years": -1, "tac_usd": 0, "lcop": -1,
            },
        },
        "tax": {
            "corporate_tax_rate": 0.21,
            "state_tax_rate": 0.05,
            "macrs_class": "7-year",
            "itc_fraction": 0.0,
            "production_credits": [
                {"usd_per_unit": 3.0, "units_per_year": 500_000, "start_year": 1, "years": 10, "name": "45V"},
            ],
            "revenue_escalation": 0.02,
            "opex_escalation": 0.025,
        },
    }
    r = client.post("/api/tea/after-tax", json=body)
    assert r.status_code == 200, r.text
    data = r.json()
    assert len(data["rows"]) == 21  # year 0 + 20 operating years
    # Year 0 is -CAPEX (ITC is zero here)
    assert data["rows"][0]["net_cf"] == pytest.approx(-10_000_000)
    # Year 1 should have non-zero PTC and depreciation
    assert data["rows"][1]["ptc"] == pytest.approx(3 * 500_000)
    assert data["rows"][1]["depreciation"] > 0


def test_run_with_lcop(client: TestClient) -> None:
    body = {
        "scenario": {
            "cepci": 800,
            "project_life_years": 20,
            "discount_rate": 0.10,
            "annual_production": 1_000_000,  # 1000 ton/yr
        },
        "blocks": [
            {
                "id": "P1",
                "type": "Pump",
                "work": 50_000,
                "inlet": {"T": 300, "P": 1e5, "flow_mol": 10},
                "outlet": {"T": 300, "P": 5e5, "flow_mol": 10},
            },
        ],
    }
    r = client.post("/api/tea/run", json=body)
    assert r.status_code == 200
    econ = r.json()["economics"]
    assert econ["lcop"] > 0
