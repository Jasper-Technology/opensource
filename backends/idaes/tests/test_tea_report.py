"""Tests for Excel report export."""
from __future__ import annotations
from io import BytesIO

from openpyxl import load_workbook

from app.tea.report import build_report


def _fixture_result() -> dict:
    return {
        "status": "ok",
        "blocks": [
            {
                "id": "P1", "type": "Pump",
                "equipment_class": "Pump", "subtype": "Centrifugal",
                "material": "CarbonSteel",
                "size": 100.0, "size_unit": "kW",
                "pressure_barg": 4.0, "Fm": 1.54, "Fp": 1.0,
                "purchased_usd": 26_000, "bare_module_usd": 103_000, "total_module_usd": 121_500,
                "in_range": True, "note": "", "error": None,
            },
            {
                "id": "H1", "type": "Heater",
                "equipment_class": "HeatExchanger", "subtype": "FixedTube",
                "material": "CarbonSteel",
                "size": 120.0, "size_unit": "m2",
                "pressure_barg": 2.0, "Fm": 1.0, "Fp": 1.0,
                "purchased_usd": 45_000, "bare_module_usd": 148_400, "total_module_usd": 175_100,
                "in_range": True, "note": "LMTD 40 K", "error": None,
            },
        ],
        "totals": {
            "capex_bare_module_usd": 251_400,
            "capex_total_module_usd": 296_600,
            "opex_annual_usd": 3_000_000,
            "opex_utilities_usd": 1_000_000,
            "opex_labor_usd": 1_800_000,
            "opex_maintenance_usd": 200_000,
            "opex_feedstock_usd": 0,
            "revenue_annual_usd": 6_000_000,
        },
        "economics": {
            "npv_usd": 12_000_000,
            "irr": 0.22,
            "payback_years": 3.1,
            "tac_usd": 3_030_000,
            "lcop": 0.28,
        },
        "warnings": [],
    }


def _fixture_scenario() -> dict:
    return {
        "cepci": 800,
        "contingency": 0.18,
        "operating_hours": 8000,
        "default_material": "CarbonSteel",
        "discount_rate": 0.10,
        "project_life_years": 20,
        "annual_revenue": 6_000_000,
        "annual_feedstock_cost": 0,
        "utility_prices": {},
        "annual_production": 1_000_000,
    }


def test_build_report_produces_valid_xlsx() -> None:
    data = build_report(
        tea_result=_fixture_result(),
        scenario=_fixture_scenario(),
        project_name="Methanol Retrofit",
    )
    assert data[:4] == b"PK\x03\x04"  # zip/xlsx magic
    wb = load_workbook(BytesIO(data))
    assert set(wb.sheetnames) == {"Summary", "CAPEX", "OPEX", "Cash Flow", "Assumptions"}


def test_summary_sheet_has_kpis() -> None:
    data = build_report(tea_result=_fixture_result(), scenario=_fixture_scenario())
    wb = load_workbook(BytesIO(data))
    ws = wb["Summary"]
    # Row 6 is NPV per the template
    labels = [ws.cell(row=r, column=1).value for r in range(5, 18)]
    assert "NPV" in labels
    assert "IRR" in labels
    assert "Payback (yr)" in labels


def test_capex_sheet_row_per_block() -> None:
    data = build_report(tea_result=_fixture_result(), scenario=_fixture_scenario())
    wb = load_workbook(BytesIO(data))
    ws = wb["CAPEX"]
    # Header + 2 blocks = 3 rows minimum
    assert ws.max_row >= 3
    ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert set(ids) == {"P1", "H1"}


def test_cash_flow_sheet_rows_match_project_life() -> None:
    data = build_report(tea_result=_fixture_result(), scenario=_fixture_scenario())
    wb = load_workbook(BytesIO(data))
    ws = wb["Cash Flow"]
    # Header + year 0 + 20 years = 22 rows
    assert ws.max_row == 22
    # Year 0 should be -capex
    assert ws.cell(row=2, column=4).value == -296_600


def test_opex_percentages_sum_to_one() -> None:
    data = build_report(tea_result=_fixture_result(), scenario=_fixture_scenario())
    wb = load_workbook(BytesIO(data))
    ws = wb["OPEX"]
    # Last row is Total; penultimate rows are categories
    totals_row = ws.max_row
    assert ws.cell(row=totals_row, column=1).value == "Total"
    assert abs(ws.cell(row=totals_row, column=3).value - 1.0) < 1e-9


def test_safe_filename_characters_in_project_name() -> None:
    data = build_report(
        tea_result=_fixture_result(),
        scenario=_fixture_scenario(),
        project_name="CH3OH / syngas (v2)",
    )
    # Name with slashes shouldn't break the build
    assert data[:4] == b"PK\x03\x04"
