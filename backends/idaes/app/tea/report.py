"""
Excel report export — takes a TEAResult + ScenarioConfig and produces a
multi-sheet XLSX for download.

Sheets:
- Summary    — KPIs + scenario config
- CAPEX      — per-block cost breakdown
- OPEX       — composition table
- Cash Flow  — year-by-year DCF schedule
- Assumptions — catalog refs, correlation citations, AACE class note

R1 produces an AACE Class 4 report (~±30% accuracy). AACE Class 3
templates ship in R2.
"""
from __future__ import annotations
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid")
_CURRENCY_FMT = '"$"#,##0'
_CURRENCY_DECIMAL_FMT = '"$"#,##0.00'
_PCT_FMT = "0.00%"
_NUMERIC_FMT = "#,##0.00"


def build_report(
    *,
    tea_result: dict,
    scenario: dict,
    project_name: str = "Jasper TEA Report",
) -> bytes:
    """Return xlsx bytes ready for HTTP download."""
    wb = Workbook()
    wb.remove(wb.active)

    _summary_sheet(wb, tea_result=tea_result, scenario=scenario, project_name=project_name)
    _capex_sheet(wb, tea_result=tea_result)
    _opex_sheet(wb, tea_result=tea_result)
    _cash_flow_sheet(wb, tea_result=tea_result, scenario=scenario)
    _assumptions_sheet(wb, scenario=scenario)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Summary ─────────────────────────────────────────────────────────────────


def _summary_sheet(wb: Workbook, *, tea_result: dict, scenario: dict, project_name: str) -> None:
    ws = wb.create_sheet("Summary")
    totals = tea_result.get("totals", {})
    econ = tea_result.get("economics", {})

    ws["A1"] = project_name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(italic=True, color="888888")
    ws["A3"] = f"Class 4 estimate (±30%) — correlations from Turton 4th ed."
    ws["A3"].font = Font(italic=True, color="888888")

    # KPIs
    ws["A5"] = "Key Performance Indicators"
    ws["A5"].font = _HEADER
    ws["A5"].fill = _HEADER_FILL

    kpis: list[tuple[str, Any, str]] = [
        ("NPV", econ.get("npv_usd", 0), _CURRENCY_FMT),
        ("IRR", econ.get("irr", 0), _PCT_FMT),
        ("Payback (yr)", econ.get("payback_years", 0), _NUMERIC_FMT),
        ("TAC", econ.get("tac_usd", 0), _CURRENCY_FMT),
        ("LCOP", econ.get("lcop", -1), _CURRENCY_DECIMAL_FMT),
        ("", "", ""),
        ("CAPEX (Total Module)", totals.get("capex_total_module_usd", 0), _CURRENCY_FMT),
        ("CAPEX (Bare Module)", totals.get("capex_bare_module_usd", 0), _CURRENCY_FMT),
        ("OPEX (annual)", totals.get("opex_annual_usd", 0), _CURRENCY_FMT),
        ("Revenue (annual)", totals.get("revenue_annual_usd", 0), _CURRENCY_FMT),
    ]
    for i, (label, value, fmt) in enumerate(kpis, start=6):
        ws.cell(row=i, column=1, value=label)
        cell = ws.cell(row=i, column=2, value=value)
        if fmt and value != "":
            cell.number_format = fmt

    # Scenario config
    ws["A18"] = "Scenario"
    ws["A18"].font = _HEADER
    ws["A18"].fill = _HEADER_FILL

    scn_rows: list[tuple[str, Any]] = [
        ("CEPCI", scenario.get("cepci", 800)),
        ("Contingency", scenario.get("contingency", 0.18)),
        ("Operating hours/yr", scenario.get("operating_hours", 8000)),
        ("Default material", scenario.get("default_material", "CarbonSteel")),
        ("Discount rate", scenario.get("discount_rate", 0.10)),
        ("Project life (yr)", scenario.get("project_life_years", 20)),
        ("Annual production (units/yr)", scenario.get("annual_production", 0) or 0),
    ]
    for i, (label, value) in enumerate(scn_rows, start=19):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)

    _autosize(ws, [1, 2])


# ── CAPEX ───────────────────────────────────────────────────────────────────


def _capex_sheet(wb: Workbook, *, tea_result: dict) -> None:
    ws = wb.create_sheet("CAPEX")
    headers = [
        "Block ID", "Type", "Equipment", "Subtype", "Material",
        "Size", "Unit", "Pressure (barg)", "Fm", "Fp",
        "Purchased (USD)", "Bare Module (USD)", "Total Module (USD)", "Note",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = _HEADER
        c.fill = _HEADER_FILL

    blocks = sorted(
        tea_result.get("blocks", []),
        key=lambda b: b.get("total_module_usd") or 0,
        reverse=True,
    )
    for b in blocks:
        row = [
            b.get("id"),
            b.get("type"),
            b.get("equipment_class") or "—",
            b.get("subtype") or "—",
            b.get("material") or "—",
            _round(b.get("size"), 2),
            b.get("size_unit") or "",
            _round(b.get("pressure_barg"), 2),
            _round(b.get("Fm"), 3),
            _round(b.get("Fp"), 3),
            b.get("purchased_usd"),
            b.get("bare_module_usd"),
            b.get("total_module_usd"),
            b.get("error") or b.get("note", ""),
        ]
        ws.append(row)

    # Currency columns
    for r in range(2, len(blocks) + 2):
        for col in (11, 12, 13):
            ws.cell(row=r, column=col).number_format = _CURRENCY_FMT

    _autosize(ws, list(range(1, len(headers) + 1)))


# ── OPEX ────────────────────────────────────────────────────────────────────


def _opex_sheet(wb: Workbook, *, tea_result: dict) -> None:
    ws = wb.create_sheet("OPEX")
    totals = tea_result.get("totals", {})
    total_opex = totals.get("opex_annual_usd", 0) or 1

    ws.append(["Category", "Annual cost (USD)", "% of OPEX"])
    for col in (1, 2, 3):
        c = ws.cell(row=1, column=col)
        c.font = _HEADER
        c.fill = _HEADER_FILL

    lines = [
        ("Feedstock", totals.get("opex_feedstock_usd", 0)),
        ("Utilities", totals.get("opex_utilities_usd", 0)),
        ("Labor", totals.get("opex_labor_usd", 0)),
        ("Maintenance", totals.get("opex_maintenance_usd", 0)),
    ]
    for label, value in lines:
        ws.append([label, value, (value or 0) / total_opex if total_opex else 0])

    ws.append(["Total", total_opex, 1.0])
    ws.cell(row=len(lines) + 2, column=1).font = Font(bold=True)

    # Formats
    for r in range(2, len(lines) + 3):
        ws.cell(row=r, column=2).number_format = _CURRENCY_FMT
        ws.cell(row=r, column=3).number_format = _PCT_FMT

    _autosize(ws, [1, 2, 3])


# ── Cash Flow ───────────────────────────────────────────────────────────────


def _cash_flow_sheet(wb: Workbook, *, tea_result: dict, scenario: dict) -> None:
    ws = wb.create_sheet("Cash Flow")
    headers = ["Year", "Revenue", "OPEX", "Net Cash Flow", "Discounted CF", "Cumulative DCF"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = _HEADER
        c.fill = _HEADER_FILL

    totals = tea_result.get("totals", {})
    capex = totals.get("capex_total_module_usd", 0)
    revenue = totals.get("revenue_annual_usd", 0)
    opex = totals.get("opex_annual_usd", 0)
    life = int(scenario.get("project_life_years", 20))
    rate = float(scenario.get("discount_rate", 0.10))

    cumulative = -capex
    ws.append([0, 0, 0, -capex, -capex, cumulative])

    for t in range(1, life + 1):
        net = revenue - opex
        dcf = net / ((1 + rate) ** t)
        cumulative += dcf
        ws.append([t, revenue, -opex, net, dcf, cumulative])

    # Currency columns
    for r in range(2, life + 3):
        for col in (2, 3, 4, 5, 6):
            ws.cell(row=r, column=col).number_format = _CURRENCY_FMT

    _autosize(ws, list(range(1, len(headers) + 1)))


# ── Assumptions ─────────────────────────────────────────────────────────────


def _assumptions_sheet(wb: Workbook, *, scenario: dict) -> None:
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "Methodology"
    ws["A1"].font = _HEADER
    ws["A1"].fill = _HEADER_FILL

    lines = [
        "• Equipment costs from Turton, Analysis, Synthesis and Design of",
        "  Chemical Processes 4th ed. — log10(Cp) = K1 + K2 log10(A) + K3 log10(A)^2,",
        "  scaled to current CEPCI.",
        "• Bare module cost C_BM = Cp · (B1 + B2·Fm·Fp); total module adds",
        "  contingency.",
        "• Material factors (Fm) and pressure correlations (Fp) sourced from",
        "  Turton Tables A.3, A.4 and related figures.",
        "• Sizing: LMTD for heat exchangers (U = 500 W/m²K default), V = F·τ",
        "  for reactors (τ = 1 h default), power = |W| for rotating equipment,",
        "  5-min liquid residence at 50% holdup for flash drums.",
        "• Utility OPEX from block duties × operating hours × tiered prices.",
        "• Labor via Turton 6-operator rule scaled by equipment count.",
        "• Maintenance 3% of installed CAPEX per Turton recommendation.",
        "• Accuracy class: AACE Class 4 (conceptual, ±30%).",
        "",
        "Limitations:",
        "• No MACRS depreciation, tax holidays, or credits (45V/45Q/CBAM).",
        "• No ESG scope 1/2/3, carbon pricing, or LCA integration.",
        "• Correlation sizes clipped within published validity ranges;",
        "  see 'Note' column on CAPEX sheet for flagged extrapolations.",
    ]
    for i, line in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 80


# ── Helpers ─────────────────────────────────────────────────────────────────


def _round(v: Any, digits: int) -> Any:
    if v is None or not isinstance(v, (int, float)):
        return v
    return round(v, digits)


def _autosize(ws, cols: list[int]) -> None:
    for col in cols:
        max_len = 8
        for cell in ws[get_column_letter(col)]:
            v = cell.value
            if v is not None:
                max_len = max(max_len, min(50, len(str(v))))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2
