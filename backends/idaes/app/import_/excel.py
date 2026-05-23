"""
Excel template parser — Jasper's simplest import path.

Template shape (4 sheets; all case-insensitive headers):

Sheet ``Components``
    name, formula, cas

Sheet ``Blocks``
    id, type, name, duty, duty_unit, work, work_unit, pressure_drop,
    pressure_drop_unit

Sheet ``Streams``
    id, from_block, from_port, to_block, to_port,
    T, T_unit, P, P_unit, flow, flow_unit, <component>, <component>, ...
    (Composition columns: one column per component name; values are mole
    fractions unless the ``composition_basis`` cell says otherwise.)

Sheet ``Economics`` (optional)
    key, value, unit
    Recognised keys: product_price, feedstock_price, location,
    operating_hours, project_life_years, discount_rate, cepci.

Every unit string is normalised to SI via ``app.units.convert``; rows
with unparseable units are skipped with a Diagnostic.
"""
from __future__ import annotations
import logging
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.units import DimensionMismatchError, UnknownUnitError, convert
from .models import (
    Diagnostic,
    ImportResult,
    ImportedBlock,
    ImportedProject,
    ImportedStream,
)

logger = logging.getLogger(__name__)

# Max file size and sheet count to prevent decompression-bomb attacks.
_MAX_BYTES = 10 * 1024 * 1024  # 10 MB
_MAX_SHEETS = 20


class ExcelImportError(ValueError):
    pass


def parse_excel(data: bytes, *, filename: str = "upload.xlsx") -> ImportResult:
    """Parse an uploaded .xlsx workbook into a Jasper ImportedProject."""
    diagnostics: list[Diagnostic] = []

    if len(data) > _MAX_BYTES:
        diagnostics.append(
            Diagnostic(
                severity="error",
                message=f"File exceeds {_MAX_BYTES // 1024 // 1024} MB limit",
                location=filename,
            )
        )
        return ImportResult(project=None, diagnostics=diagnostics)

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a grab-bag of errors
        diagnostics.append(
            Diagnostic(severity="error", message=f"Cannot read workbook: {exc}", location=filename)
        )
        return ImportResult(project=None, diagnostics=diagnostics)

    if len(wb.sheetnames) > _MAX_SHEETS:
        diagnostics.append(
            Diagnostic(
                severity="error",
                message=f"Workbook has too many sheets ({len(wb.sheetnames)} > {_MAX_SHEETS})",
                location=filename,
            )
        )
        return ImportResult(project=None, diagnostics=diagnostics)

    sheets = {n.lower(): n for n in wb.sheetnames}

    components = _parse_components(wb, sheets, diagnostics)
    blocks = _parse_blocks(wb, sheets, diagnostics)
    streams = _parse_streams(wb, sheets, diagnostics, components=components)
    _parse_economics(wb, sheets, diagnostics)  # currently only records diagnostics / TODO return

    if not blocks and not streams:
        diagnostics.append(
            Diagnostic(
                severity="error",
                message="Workbook has no blocks or streams — is this the right template?",
                location=filename,
            )
        )
        return ImportResult(project=None, diagnostics=diagnostics)

    project_name = filename.rsplit(".", 1)[0]
    return ImportResult(
        project=ImportedProject(
            name=project_name,
            components=components,
            blocks=blocks,
            streams=streams,
        ),
        diagnostics=diagnostics,
    )


# ── Sheet parsers ───────────────────────────────────────────────────────────


def _header_row(ws: Worksheet) -> Optional[dict[str, int]]:
    """Read the first non-empty row as column headers. Returns ``{name: col_index}``."""
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            return None
        return {
            str(v).strip().lower(): idx
            for idx, v in enumerate(row)
            if v is not None and str(v).strip()
        }
    return None


def _cell(row: tuple[Any, ...], col: Optional[int]) -> Any:
    if col is None or col >= len(row):
        return None
    return row[col]


def _parse_components(wb, sheets: dict[str, str], diagnostics: list[Diagnostic]) -> list[str]:
    if "components" not in sheets:
        return []
    ws: Worksheet = wb[sheets["components"]]
    header = _header_row(ws)
    if not header or "name" not in header:
        diagnostics.append(
            Diagnostic(
                severity="warning",
                message="Components sheet missing 'name' header — skipping",
                location="Components",
            )
        )
        return []

    out: list[str] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = _cell(row, header.get("name"))
        if name is None:
            continue
        out.append(str(name).strip())
    return out


def _parse_blocks(wb, sheets: dict[str, str], diagnostics: list[Diagnostic]) -> list[ImportedBlock]:
    if "blocks" not in sheets:
        diagnostics.append(
            Diagnostic(
                severity="error",
                message="Missing required 'Blocks' sheet",
                location="Blocks",
            )
        )
        return []
    ws: Worksheet = wb[sheets["blocks"]]
    header = _header_row(ws)
    if not header or "id" not in header or "type" not in header:
        diagnostics.append(
            Diagnostic(
                severity="error",
                message="Blocks sheet must have at least 'id' and 'type' columns",
                location="Blocks",
            )
        )
        return []

    blocks: list[ImportedBlock] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        block_id = _cell(row, header.get("id"))
        block_type = _cell(row, header.get("type"))
        if block_id is None or block_type is None:
            continue

        name = _cell(row, header.get("name")) or str(block_id)
        duty = _quantity_to_si(
            row, header, "duty", "duty_unit", "W",
            f"Blocks!row{r_idx}", diagnostics,
        )
        work = _quantity_to_si(
            row, header, "work", "work_unit", "W",
            f"Blocks!row{r_idx}", diagnostics,
        )
        pdrop = _quantity_to_si(
            row, header, "pressure_drop", "pressure_drop_unit", "Pa",
            f"Blocks!row{r_idx}", diagnostics,
        )

        blocks.append(
            ImportedBlock(
                id=str(block_id).strip(),
                type=str(block_type).strip(),
                name=str(name).strip(),
                duty=duty,
                work=work,
                pressure_drop=pdrop,
            )
        )
    return blocks


def _parse_streams(
    wb,
    sheets: dict[str, str],
    diagnostics: list[Diagnostic],
    components: list[str],
) -> list[ImportedStream]:
    if "streams" not in sheets:
        return []
    ws: Worksheet = wb[sheets["streams"]]
    header = _header_row(ws)
    if not header or "id" not in header:
        diagnostics.append(
            Diagnostic(
                severity="error",
                message="Streams sheet missing 'id' column",
                location="Streams",
            )
        )
        return []

    # Component composition columns — any header that matches a component name
    comp_lower = {c.lower(): c for c in components}
    comp_cols: dict[str, int] = {
        comp_lower[name]: idx for name, idx in header.items() if name in comp_lower
    }

    streams: list[ImportedStream] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        stream_id = _cell(row, header.get("id"))
        if stream_id is None:
            continue

        T = _quantity_to_si(row, header, "t", "t_unit", "K", f"Streams!row{r_idx}", diagnostics)
        P = _quantity_to_si(row, header, "p", "p_unit", "Pa", f"Streams!row{r_idx}", diagnostics)
        # Flow: accept any flow unit; SI target depends on dimension. Default to molar.
        flow_val, flow_mol, flow_mass = _parse_stream_flow(row, header, f"Streams!row{r_idx}", diagnostics)

        comp: dict[str, float] = {}
        for cname, cidx in comp_cols.items():
            v = _cell(row, cidx)
            if v is None:
                continue
            try:
                comp[cname] = float(v)
            except (TypeError, ValueError):
                diagnostics.append(
                    Diagnostic(
                        severity="warning",
                        message=f"Composition cell not numeric: {v!r}",
                        location=f"Streams!row{r_idx}!{cname}",
                    )
                )

        streams.append(
            ImportedStream(
                id=str(stream_id).strip(),
                from_block=_str_or_none(_cell(row, header.get("from_block"))),
                from_port=_str_or_none(_cell(row, header.get("from_port"))),
                to_block=_str_or_none(_cell(row, header.get("to_block"))),
                to_port=_str_or_none(_cell(row, header.get("to_port"))),
                T=T,
                P=P,
                flow_mol=flow_mol,
                flow_mass=flow_mass,
                composition=comp,
            )
        )
    return streams


def _parse_economics(wb, sheets: dict[str, str], diagnostics: list[Diagnostic]) -> None:
    """Placeholder — parse key/value pairs into a ScenarioConfig (TODO wire into result)."""
    if "economics" not in sheets:
        return
    # Currently just validates the sheet shape; values will be wired into
    # the ImportResult in a follow-up commit.
    ws: Worksheet = wb[sheets["economics"]]
    header = _header_row(ws)
    if not header or "key" not in header or "value" not in header:
        diagnostics.append(
            Diagnostic(
                severity="warning",
                message="Economics sheet missing 'key'/'value' columns — values ignored",
                location="Economics",
            )
        )


# ── Helpers ─────────────────────────────────────────────────────────────────


def _str_or_none(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _quantity_to_si(
    row: tuple[Any, ...],
    header: dict[str, int],
    value_key: str,
    unit_key: str,
    si_target: str,
    where: str,
    diagnostics: list[Diagnostic],
) -> Optional[float]:
    """Pull ``value_key`` + ``unit_key`` cells and convert to the SI target."""
    value = _cell(row, header.get(value_key))
    unit = _cell(row, header.get(unit_key))
    if value is None or value == "":
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        diagnostics.append(
            Diagnostic(
                severity="warning",
                message=f"Cell {value_key!r} is not numeric: {value!r}",
                location=where,
            )
        )
        return None
    if unit is None or unit == "":
        # Assume SI if unit omitted
        return v
    try:
        return convert(v, str(unit), si_target)
    except (UnknownUnitError, DimensionMismatchError) as exc:
        diagnostics.append(
            Diagnostic(severity="warning", message=str(exc), location=where)
        )
        return None


def _parse_stream_flow(
    row: tuple[Any, ...],
    header: dict[str, int],
    where: str,
    diagnostics: list[Diagnostic],
) -> tuple[Optional[float], Optional[float], Optional[float]]:
    """Flow column can be molar or mass; dispatch on the unit string."""
    value = _cell(row, header.get("flow"))
    unit = _cell(row, header.get("flow_unit"))
    if value is None or value == "":
        return None, None, None
    try:
        v = float(value)
    except (TypeError, ValueError):
        diagnostics.append(
            Diagnostic(severity="warning", message=f"Flow cell not numeric: {value!r}", location=where)
        )
        return None, None, None

    unit_s = str(unit) if unit else "mol/s"
    # Try molar first
    for target, mol, mass in [("mol/s", True, False), ("kg/s", False, True)]:
        try:
            si = convert(v, unit_s, target)
            return (v, si, None) if mol else (v, None, si)
        except DimensionMismatchError:
            continue
        except UnknownUnitError as exc:
            diagnostics.append(
                Diagnostic(severity="warning", message=str(exc), location=where)
            )
            return None, None, None
    diagnostics.append(
        Diagnostic(
            severity="warning",
            message=f"Flow unit {unit_s!r} is neither molar nor mass; ignored",
            location=where,
        )
    )
    return None, None, None
